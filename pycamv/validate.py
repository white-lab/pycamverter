"""
This module provides functionality for interacting with CAMV.

Currently limited to importing and outputing scan lists.
"""

# Built-ins
from __future__ import absolute_import, division

from collections import OrderedDict
from functools import partial
import logging
import multiprocessing
import re
import tempfile

from openpyxl import load_workbook

from . import compare, fragments, gen_sequences, mascot, ms_labels, scans
from .utils import LenGen


LOGGER = logging.getLogger("pycamv.validate")


class SearchOptions:
    """
    Contains options used to search a data set in MASCOT.

    Attributes
    ----------
    label_type : tuple of (str, int)
    """
    def __init__(self, fixed_mods, var_mods):
        self.label_type = (None, 0)

        for mod in fixed_mods:
            mod = re.sub(r"([\w-]+) \([\w-]+\)", r"\1", mod)
            num = ms_labels.LABEL_NUMBERS.get(mod, 0)

            if num > 0:
                self.label_type = ("Fixed", num)

        for mod in var_mods:
            mod = re.sub(r"([\w-]+) \([\w-]+\)", r"\1", mod)
            num = ms_labels.LABEL_NUMBERS.get(mod, 0)

            if num > 0:
                self.label_type = ("Variable", num)

        # TODO: Parse out SILAC, C-mod, phospho, etc


def _remap_pst(pep_mods):
    return [
        (
            count,
            mod,
            letters + (
                ("Y",)
                if (mod == "Phospho" and set(letters) == set(["S", "T"])) else
                ()
            ),
        )
        for count, mod, letters in pep_mods
    ]


def load_scan_list(scans_path):
    wb = load_workbook(scans_path)
    ws = wb.active
    return [row[0].value for row in ws.iter_rows()]


def _map_seq(kv):
    pep_query = kv
    return (
        pep_query,
        tuple(
            gen_sequences.gen_possible_seq(
                pep_query.pep_seq,
                pep_query.pep_mods,
            )
        ),
    )


def _map_frag(kv):
    pep_query, sequence = kv
    return (
        (pep_query, tuple(sequence)),
        fragments.fragment_ions(
            sequence, pep_query.pep_exp_z,
        ),
    )


def _map_compare(kv, scan_mapping):
    key, val = kv
    pep_query, sequence = key
    frag_ions, scan = val

    peaks = compare.compare_spectra(
        scan,
        frag_ions,
        pep_query.pep_exp_z,
        scans.c13_num(pep_query, scan_mapping[pep_query]),
        tol=compare.COLLISION_TOLS[scan_mapping[pep_query].collision_type],
    )

    del scan

    return key, peaks


def validate_spectra(
    xml_path,
    raw_path,
    scans_path=None,
    scan_list=None,
    cpu_count=None,
):
    """
    Generate CAMV web page for validating spectra.

    Parameters
    ----------
    xml_path : str
    raw_path : str
    scans_path : str, optional
    scan_list : list of int, optional
    cpu_count : int, optional

    Returns
    -------
    options : :class:`pycamv.validate.SearchOptions`
    peak_hits : dict of (tuple of :class:`pycamv.mascot.PeptideQuery`, list),
    list of :class:`pycamv.compare.PeptideHit`
        Dictionary mapping peptide queries and their sequences to peptide hits.
    scan_mapping : OrderedDict of :class:`pycamv.mascot.PeptideQuery`,
    :class:`pycamv.scans.ScanQuery`
        Dictionary mapping peptide queries to their scan data, including
        collision type and size of isolation window.
    precursor_windows : dict of :class:`pycamv.mascot.PeptideQuery`, list
        Dictionary mapping peptide queries to peak lists for precursor scans.
    label_windows : dict of :class:`pycamv.mascot.PeptideQuery`, list
        Dictionary mapping peptide queries to peak lists for quantification
        channels.
    """
    if cpu_count is None:
        try:
            cpu_count = multiprocessing.cpu_count()
        except NotImplementedError:
            cpu_count = 2

    LOGGER.debug("Validating data set with {} cpus".format(cpu_count))

    if scan_list is None:
        scan_list = []

    if scans_path is not None:
        scan_list += load_scan_list(scans_path)

    # Read MASCOT xml file
    fixed_mods, var_mods, pep_queries = mascot.read_mascot_xml(xml_path)

    # Optionally filter queries using a scan list
    if scan_list:
        pep_queries = [
            query
            for query in pep_queries
            if query.scan in scan_list
        ]

    # Remap pST -> pSTY
    LOGGER.info("Remapping pST -> pSTY")
    for pep_query in pep_queries:
        pep_query.pep_var_mods = _remap_pst(pep_query.pep_var_mods)
        pep_query.pep_fixed_mods = _remap_pst(pep_query.pep_fixed_mods)

    # Extract MASCOT search options
    options = SearchOptions(fixed_mods, var_mods)

    # Remove peptides with an excess of modification combinations

    # Get scan data from RAW file
    out_dir = tempfile.mkdtemp()

    LOGGER.info("Getting scan data.")

    scan_queries, ms_two_data, ms_data = scans.get_scan_data(
        raw_path, pep_queries, out_dir,
    )

    scan_mapping = OrderedDict(
        (pep_query, scan_query)
        for pep_query, scan_query in zip(pep_queries, scan_queries)
    )

    # Generate sequences
    LOGGER.info(
        "Generating all possible sequence-modification combinations."
    )
    pool = multiprocessing.Pool(processes=cpu_count)
    sequence_mapping = OrderedDict(
        pool.map(_map_seq, pep_queries)
    )

    LOGGER.info(
        "Generating fragment ions for {} queries.".format(
            len(sequence_mapping),
        )
    )
    fragment_mapping = OrderedDict(
        pool.map(
            _map_frag,
            (
                (pep_query, sequence)
                for pep_query, sequences in sequence_mapping.items()
                for sequence in sequences
            ),
        )
    )

    LOGGER.info(
        "Comparing predicted to actual peaks for {} spectra.".format(
            len(fragment_mapping),
        )
    )

    def _imap_compare(kv):
        key, val = kv
        return (key, (val, ms_two_data[key[0].scan].deRef()))

    peak_hits = dict(
        pool.map(
            partial(_map_compare, scan_mapping=scan_mapping),
            LenGen(
                (
                    (key, (val, ms_two_data[key[0].scan].deRef()))
                    for key, val in fragment_mapping.items()
                ),
                len(fragment_mapping),
            ),
            cpu_count * 4,
        )
    )

    del pool

    # XXX: Determine SILAC precursor masses?

    LOGGER.info("Collecting precursor ion peaks.")
    precursor_windows = OrderedDict(
        zip(
            pep_queries,
            scans.get_precursor_peak_window(scan_queries, ms_data)
        )
    )

    LOGGER.info("Collecting peptide label peaks.")
    label_windows = OrderedDict(
        zip(
            pep_queries,
            scans.get_label_peak_window(pep_queries, ms_two_data)
        )
    )

    # XXX: Remove precursor contaminated scans from validation list?

    # Check each assignment to each scan

    # Output data

    # shutil.rmtree(out_dir)

    return options, peak_hits, scan_mapping, precursor_windows, label_windows
