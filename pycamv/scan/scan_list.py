
from __future__ import absolute_import, division

import logging
import os

LOGGER = logging.getLogger("pycamv.scan_list")


def load_scan_xlsx(path):
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb.active
    rows = [row[0].value for row in ws.iter_rows()]
    rows = [i for i in rows if isinstance(i, int)]

    return rows


def load_scan_csv(path):
    with open(path) as f:
        rows = [
            i.strip()
            for line in f
            for i in line.split(",")
        ]

    rows = [int(i) for i in rows if i.isdigit()]

    return rows


BACKENDS = {
    ".csv": load_scan_csv,
    ".xlsx": load_scan_xlsx,
}


def load_scan_list(path):
    ext = os.path.splitext(path)[1]
    backend = BACKENDS.get(ext)

    if backend is None:
        raise Exception("Unable to open scan list: {}".format(path))

    LOGGER.debug("Using {} backend for {}".format(backend.__name__, ext))

    return backend(path)
